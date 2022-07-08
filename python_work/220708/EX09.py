from openpyxl import load_workbook

wb = load_workbook('c.xlsx')
ws = wb.active

ws['A11'] = "=SUM(A1:A10)"
#한줄 삽입
ws.append([11,22,33,44,55,66,77,88,99])

# for cell in ws.insert_rows:
#     for 


wb.save('c.xlsx')
wb.close()